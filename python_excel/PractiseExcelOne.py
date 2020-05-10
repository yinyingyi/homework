'''
使用openpyxl实现以下需求
使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
'''

from openpyxl import Workbook, load_workbook

class Practice():       #定义一个练习类
    def __init__(self):     #定义data方法，用来存放待写入数据
        self.name = ["小红","小华","李雷","韩梅梅"]      #定义姓名列表
        self.data1 = {160:51,179:71,180:74,158:44}      #定义身高体重字典
        self.data1_keys = [i for i in self.data1.keys()]     #使用列表获取字典key值。要想单独获取字典每个值，必须先转化为列表

    def creat_excel_one(self):       #创建一个名字为excel_one的.xlsx文件
        wb = Workbook()     # 实例化Workbook()，alt+enter快捷导入
        ws1 = wb.active         # 获取当前页签
        ws1.title = "one"       # 给当前页签命名one
        ws1["A1"] = "姓名"    # 在A1的位置输入"姓名"
        ws1["B1"] = "身高"    # 在B1的位置输入"身高"
        ws1["C1"] = "体重"    # 在C1的位置输入"体重"
        ws1["D1"] = "是否超重"      # 在D1的位置输入"是否超重"

        for i in range(len(self.name)):     #for循环，循环次数i=name长度
            ws1.cell(row=i + 2, column=1).value = self.name[i]      #第一列name。row行，col列，从excle第二行开始写值，所以row=i + 2
            ws1.cell(row=i + 2, column=2).value = self.data1_keys[i]       #第二列height。获取的字典已经转换好的key值
            ws1.cell(row=i + 2, column=3).value = self.data1[self.data1_keys[i]]      #第三列weight。通过key值获取对应的value

        wb.save("excel_one.xlsx")       #保存excel_one.xlsx数据

    def health_data(self):      #定义health_data方法
        ld = load_workbook(filename="excel_one.xlsx")       #读取excel_one.xlsx文件
        sheet = ld["one"]   #指定sheet页
        for i in range(4):      #循环获取cell里的值。共4组数据，所以range(4)
            height = sheet.cell(row=i + 2, column=2).value      #获取height值
            weight = sheet.cell(row=i + 2, column=3).value      #获取weight值
            heath_weight = (height - 70) *0.6       #定义健康体重计算公式
            if weight > heath_weight:       #如果>健康体重，输出超标，并且在文档第4列标注
                print("{}超标".format(self.name[i]))
                sheet.cell(row=i + 2, column=4).value="超标"
            elif weight == heath_weight:        #如果=健康体重，输出标准，并且在文档第4列标注
                print("{}标准".format(self.name[i]))
                sheet.cell(row=i + 2, column=4).value = "标准"
            else:       #如果>健康体重，输出太瘦，并且在文档第4列标注
                print("{}太瘦".format(self.name[i]))
                sheet.cell(row=i + 2, column=4).value = "太瘦"
            ld.save("excel_one.xlsx")

p = Practice()      #初始化Practice()类
p.health_data()     #调用health_data()方法